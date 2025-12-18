# File: backend/msp/management/commands/import_edgar.py
import csv
import openpyxl
import os
from django.core.management.base import BaseCommand
from django.db import transaction
import pandas as pd

from msp.models import RawFirm, CandidateFirm
from msp.utils import evaluate_raw_row
from msp.tasks import evaluate_and_promote

class Command(BaseCommand):
    help = 'Import EDGAR/Excel data into RawFirm rows and optionally auto-filter to CandidateFirm'

    def add_arguments(self, parser):
        parser.add_argument('path', type=str)
        parser.add_argument('--source', type=str, default='EDGAR')
        parser.add_argument('--auto-filter', action='store_true')
        parser.add_argument('--threshold', type=float, default=50.0)
        parser.add_argument('--sheet', type=str, default=None)
        parser.add_argument('--dry-run', action='store_true')
        parser.add_argument('--background', action='store_true', help='Schedule evaluation tasks instead of immediate processing')
        parser.add_argument('--chunk-size', type=int, default=1000, help='Rows per chunk (csv only)')

    def handle(self, *args, **options):
        path = options['path']
        source = options['source']
        auto_filter = options['auto_filter']
        threshold = options['threshold']
        sheet = options['sheet']
        dry_run = options['dry_run']
        background = options['background']
        chunk_size = options['chunk_size']

        self.stdout.write(f'Reading {path}...')
        created = promoted = 0

        lower = path.lower()
        # CSV chunked
        if lower.endswith('.csv'):
            with open(path, newline='', encoding='utf-8') as fh:
                reader = csv.DictReader(fh)
                buffer = []
                for row in reader:
                    buffer.append(row)
                    if len(buffer) >= chunk_size:
                        created_chunk, promoted_chunk = self._process_rows(buffer, source, auto_filter, threshold, dry_run, background)
                        created += created_chunk; promoted += promoted_chunk
                        buffer = []
                if buffer:
                    created_chunk, promoted_chunk = self._process_rows(buffer, source, auto_filter, threshold, dry_run, background)
                    created += created_chunk; promoted += promoted_chunk
        else:
            # For xlsx: use openpyxl streaming or pandas if small
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheetnames = wb.sheetnames
            ws = wb[sheet] if sheet and sheet in sheetnames else wb[sheetnames[0]]
            headers = None
            rows_buffer = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c).strip() if c is not None else '' for c in row]
                    continue
                payload = {headers[idx]: (row[idx] if idx < len(row) else '') for idx in range(len(headers))}
                rows_buffer.append(payload)
                if len(rows_buffer) >= chunk_size:
                    c,p = self._process_rows(rows_buffer, source, auto_filter, threshold, dry_run, background)
                    created += c; promoted += p
                    rows_buffer = []
            if rows_buffer:
                c,p = self._process_rows(rows_buffer, source, auto_filter, threshold, dry_run, background)
                created += c; promoted += p

        self.stdout.write(self.style.SUCCESS(f'Imported {created} rows. Promoted {promoted} candidates.'))

    def _normalize_company_name(self, v):
        return (v or '').strip()

    def _process_rows(self, rows, source, auto_filter, threshold, dry_run, background):
        created = promoted = 0
        with transaction.atomic():
            for idx, payload in enumerate(rows):
                company_name = (payload.get('company_name') or payload.get('Company') or payload.get('Company Name') or payload.get('company') or payload.get('company name') or '').strip()
                if not company_name:
                    continue
                if dry_run:
                    created += 1
                    if auto_filter:
                        # evaluate locally
                        fake_raw = type('F', (), {'raw_payload': payload, 'country': payload.get('country','') , 'website': payload.get('website','')})
                        score, matched = evaluate_raw_row(fake_raw) 
                        if score >= threshold:
                            promoted += 1
                    continue

                raw, created_flag = RawFirm.objects.get_or_create(
                    source=source,
                    source_id=str(payload.get('id') or payload.get('ID') or payload.get('cik') or f'{company_name}_{idx}'),
                    defaults={
                        'company_name': company_name,
                        'website': payload.get('website', ''),
                        'phone': payload.get('phone', ''),
                        'email': payload.get('email', ''),
                        'country': payload.get('country', ''),
                        'state': payload.get('state', ''),
                        'city': payload.get('city', ''),
                        'raw_payload': payload,
                    }
                )
                if created_flag:
                    created += 1

                if auto_filter:
                    if background:
                        # queue task to evaluate and possibly promote
                        evaluate_and_promote.delay(raw.id, threshold)
                    else:
                        score, matched = evaluate_raw_row(raw)
                        if score >= threshold:
                            CandidateFirm.objects.get_or_create(
                                raw_firm=raw,
                                defaults={
                                    'score': score,
                                    'matched_rules': matched,
                                    'is_us': 'us' in matched,
                                    'suspected_msp': 'msp_keyword' in matched,
                                    'suspected_pe': 'pe_keyword' in matched,
                                },
                            )
                            promoted += 1
        return created, promoted
